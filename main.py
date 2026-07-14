from __future__ import annotations

import hashlib
import hmac
import io
import json
import os
import re
import secrets
import threading
import time
import tempfile
import zipfile
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import httpx
from fastapi import Depends, FastAPI, Header, HTTPException, Request
from fastapi.responses import HTMLResponse, Response
from openpyxl import load_workbook


API_KEY = os.getenv("ACTION_API_KEY", "").strip()
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "25"))
PUBLIC_BASE_URL = os.getenv(
    "PUBLIC_BASE_URL",
    "https://powerbi-action-clean-public.onrender.com",
).rstrip("/")
DOWNLOAD_TTL_SECONDS = int(os.getenv("DOWNLOAD_TTL_SECONDS", "900"))

_DOWNLOADS: dict[str, dict[str, Any]] = {}
_DOWNLOADS_LOCK = threading.Lock()


def _prune_downloads() -> None:
    now = time.time()
    expired = [
        token
        for token, item in _DOWNLOADS.items()
        if float(item["expires_at"]) <= now
    ]
    for token in expired:
        _DOWNLOADS.pop(token, None)


def store_download(name: str, content: bytes) -> str:
    token = secrets.token_urlsafe(32)
    with _DOWNLOADS_LOCK:
        _prune_downloads()
        _DOWNLOADS[token] = {
            "name": name,
            "content": content,
            "expires_at": time.time() + DOWNLOAD_TTL_SECONDS,
        }
    return f"{PUBLIC_BASE_URL}/v1/downloads/{token}"


ALLOWED_FILE_HOST_SUFFIXES = tuple(
    value.strip().lower()
    for value in os.getenv(
        "ALLOWED_FILE_HOST_SUFFIXES",
        "oaiusercontent.com,openai.com",
    ).split(",")
    if value.strip()
)

app = FastAPI(
    title="Power BI Automation Pilot API",
    version="1.2.0",
    description=(
        "Recibe un Excel final, valida su estructura y devuelve un paquete técnico ZIP "
        "para Power BI sin modificar el archivo fuente."
    ),
)


def require_api_key(authorization: str | None = Header(default=None)) -> None:
    if not API_KEY:
        raise HTTPException(status_code=500, detail="ACTION_API_KEY no está configurada.")
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Falta autenticación Bearer.")
    supplied = authorization.removeprefix("Bearer ").strip()
    if not hmac.compare_digest(supplied, API_KEY):
        raise HTTPException(status_code=403, detail="Clave de API inválida.")


def normalize_file_ref(raw_ref: Any) -> dict[str, Any]:
    if isinstance(raw_ref, dict):
        return raw_ref
    if isinstance(raw_ref, str):
        try:
            parsed = json.loads(raw_ref)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass
    raise HTTPException(status_code=422, detail="Referencia de archivo inválida.")


def validate_download_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme != "https" or not parsed.hostname:
        raise HTTPException(status_code=422, detail="El archivo requiere una URL HTTPS válida.")
    host = parsed.hostname.lower().rstrip(".")
    allowed = any(
        host == suffix or host.endswith("." + suffix)
        for suffix in ALLOWED_FILE_HOST_SUFFIXES
    )
    if not allowed:
        raise HTTPException(status_code=422, detail="Host de archivo no autorizado.")


async def download_excel(raw_ref: Any) -> tuple[bytes, dict[str, Any]]:
    ref = normalize_file_ref(raw_ref)
    name = Path(str(ref.get("name") or "source.xlsx")).name
    url = str(ref.get("download_link") or "")
    mime_type = str(ref.get("mime_type") or "")
    file_id = str(ref.get("id") or "")

    if not name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Solo se aceptan archivos .xlsx.")
    validate_download_url(url)

    max_bytes = MAX_UPLOAD_MB * 1024 * 1024
    total = 0
    chunks: list[bytes] = []

    async with httpx.AsyncClient(timeout=45.0, follow_redirects=False) as client:
        async with client.stream("GET", url) as response:
            if response.status_code != 200:
                raise HTTPException(
                    status_code=502,
                    detail=f"No se pudo descargar el archivo temporal: HTTP {response.status_code}.",
                )
            async for chunk in response.aiter_bytes():
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(
                        status_code=413,
                        detail="El archivo supera el límite permitido.",
                    )
                chunks.append(chunk)

    data = b"".join(chunks)
    if not data:
        raise HTTPException(status_code=422, detail="El archivo recibido está vacío.")
    if not zipfile.is_zipfile(io.BytesIO(data)):
        raise HTTPException(status_code=422, detail="El archivo no es un XLSX válido.")

    return data, {
        "file_name": name,
        "file_id": file_id,
        "mime_type": mime_type,
        "size_bytes": len(data),
        "sha256": hashlib.sha256(data).hexdigest(),
        "source_modified": False,
    }


def profile_excel(data: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        bad_member = archive.testzip()
        if bad_member is not None:
            raise HTTPException(
                status_code=422,
                detail=f"El XLSX contiene un componente dañado: {bad_member}",
            )

    with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp:
        temp.write(data)
        temp.flush()
        try:
            workbook = load_workbook(
                temp.name,
                read_only=False,
                data_only=False,
                keep_links=True,
            )
        except Exception as exc:
            raise HTTPException(
                status_code=422,
                detail=f"No se pudo abrir la estructura del Excel: {exc}",
            ) from exc

        sheets: list[dict[str, Any]] = []
        tables: list[dict[str, Any]] = []
        formula_count = 0

        for worksheet in workbook.worksheets:
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "table_count": len(worksheet.tables),
                }
            )
            for table in worksheet.tables.values():
                tables.append(
                    {
                        "sheet": worksheet.title,
                        "name": table.name,
                        "ref": table.ref,
                    }
                )
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formula_count += 1

        workbook.close()

    return {
        "workbook_opened": True,
        "ooxml_zip_integrity": "passed",
        "sheet_count": len(sheets),
        "sheets": sheets,
        "table_count": len(tables),
        "tables": tables,
        "formula_count": formula_count,
        "native_excel_validation": "not_performed",
    }


def safe_package_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip()).strip("._")
    return (cleaned or "PowerBI_Package")[:100]


def create_package(
    project_name: str,
    source_meta: dict[str, Any],
    source_profile: dict[str, Any],
    approved_stages: list[int],
    specification: dict[str, Any],
) -> tuple[str, bytes, str]:
    package_name = safe_package_name(project_name) + ".zip"

    manifest = {
        "project_name": project_name,
        "source_file": source_meta["file_name"],
        "source_sha256": source_meta["sha256"],
        "source_modified": False,
        "approved_stages": approved_stages,
        "generation_status": "package_ready",
        "native_validation": "not_performed",
        "result_type": "technical_package_only",
        "pbix_generated": False,
        "pbip_generated": False,
    }

    validation_report = {
        "package_structure": "passed",
        "source_modified": False,
        "source_profile_created": True,
        "native_power_bi_validation": "not_performed",
        "notes": [
            "El paquete contiene especificaciones técnicas aprobadas.",
            "No constituye un PBIX ni un PBIP.",
        ],
    }

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        archive.writestr(
            "source_profile.json",
            json.dumps(source_profile, ensure_ascii=False, indent=2),
        )
        archive.writestr(
            "power_query.json",
            json.dumps(specification["power_query"], ensure_ascii=False, indent=2),
        )
        archive.writestr(
            "relationships.json",
            json.dumps(specification["relationships"], ensure_ascii=False, indent=2),
        )
        archive.writestr("measures.dax", specification["measures_dax"])
        archive.writestr(
            "visual_spec.json",
            json.dumps(specification["visual_spec"], ensure_ascii=False, indent=2),
        )
        archive.writestr(
            "test_plan.json",
            json.dumps(specification["test_plan"], ensure_ascii=False, indent=2),
        )
        archive.writestr(
            "validation_report.json",
            json.dumps(validation_report, ensure_ascii=False, indent=2),
        )

    content = buffer.getvalue()
    return package_name, content, hashlib.sha256(content).hexdigest()


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "service": "powerbi-action-clean"}


@app.get("/privacy", response_class=HTMLResponse)
def privacy() -> str:
    return """
    <!doctype html>
    <html lang="es">
      <head><meta charset="utf-8"><title>Política de privacidad</title></head>
      <body style="font-family:system-ui;max-width:760px;margin:40px auto;line-height:1.5">
        <h1>Política de privacidad del piloto</h1>
        <p>Este servicio procesa temporalmente archivos Excel para crear un perfil técnico
        y un paquete ZIP solicitado por el usuario.</p>
        <p>El Excel se trata como solo lectura, no se incluye en el ZIP generado y no se
        modifica. No se deben cargar credenciales, contraseñas ni información sensible.</p>
      </body>
    </html>
    """


@app.get("/v1/downloads/{token}", include_in_schema=False)
def download_package(token: str) -> Response:
    with _DOWNLOADS_LOCK:
        _prune_downloads()
        item = _DOWNLOADS.get(token)

    if item is None:
        raise HTTPException(
            status_code=404,
            detail="El archivo temporal no existe o ya venció.",
        )

    return Response(
        content=bytes(item["content"]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{item["name"]}"',
            "Cache-Control": "no-store",
            "X-Content-Type-Options": "nosniff",
        },
    )


@app.post(
    "/v1/build-package",
    operation_id="buildPowerBIPackage",
    dependencies=[Depends(require_api_key)],
)
async def build_power_bi_package(request: Request) -> dict[str, Any]:
    body = await request.json()

    project_name = str(body.get("project_name", "")).strip()
    approved_stages = body.get("approved_stages", [])
    refs = body.get("openaiFileIdRefs")
    specification = body.get("specification")

    if not project_name:
        raise HTTPException(status_code=422, detail="project_name es obligatorio.")
    if not isinstance(approved_stages, list):
        raise HTTPException(status_code=422, detail="approved_stages debe ser una lista.")
    approved_stages = sorted(
        {
            int(stage)
            for stage in approved_stages
            if isinstance(stage, int) and 1 <= stage <= 7
        }
    )
    if approved_stages != [1, 2, 3, 4, 5, 6, 7]:
        raise HTTPException(
            status_code=409,
            detail="Las etapas 1 a 7 deben estar aprobadas antes de generar el paquete.",
        )
    if not isinstance(refs, list) or len(refs) != 1:
        raise HTTPException(
            status_code=422,
            detail="Debe enviarse exactamente un Excel mediante openaiFileIdRefs.",
        )
    if not isinstance(specification, dict):
        raise HTTPException(status_code=422, detail="specification es obligatoria.")

    required = {
        "power_query",
        "relationships",
        "measures_dax",
        "visual_spec",
        "test_plan",
    }
    missing = sorted(required - set(specification))
    if missing:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Faltan componentes de la especificación.",
                "missing": missing,
            },
        )

    excel_bytes, source_meta = await download_excel(refs[0])
    source_profile = profile_excel(excel_bytes)
    package_name, package_bytes, package_sha256 = create_package(
        project_name=project_name,
        source_meta=source_meta,
        source_profile=source_profile,
        approved_stages=approved_stages,
        specification=specification,
    )

    return {
        "status": "package_ready",
        "source_modified": False,
        "native_validation": "not_performed",
        "artifact": {
            "name": package_name,
            "size_bytes": len(package_bytes),
            "sha256": package_sha256,
        },
        "source_profile": {
            "sheet_count": source_profile["sheet_count"],
            "table_count": source_profile["table_count"],
            "formula_count": source_profile["formula_count"],
        },
        "download_url": store_download(package_name, package_bytes),
        "download_expires_seconds": DOWNLOAD_TTL_SECONDS,
    }
